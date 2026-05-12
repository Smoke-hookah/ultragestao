"""Leitor de planilhas Excel."""

import openpyxl
from pathlib import Path
from typing import List, Dict, Optional, Iterator
import re
import time
from utils.logger import logger
from models.alocacao import Alocacao


class LeitorPlanilha:
    """Lê dados de planilhas de alocação."""
    
    def __init__(self, caminho_arquivo: str):
        """
        Inicializa o leitor.
        
        Args:
            caminho_arquivo: Caminho da planilha
        """
        self.caminho = Path(caminho_arquivo)
        self.nome_arquivo = self.caminho.stem
        
        if not self.caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho_arquivo}")
        
        if not self.caminho.suffix.lower() in ['.xlsx', '.xls']:
            raise ValueError("Arquivo deve ser .xlsx ou .xls")
    
    def iter_dados_brutos(self) -> Iterator[Dict]:
        """Itera pelos dados brutos da planilha (streaming, mais rápido e com menos memória)."""
        t0 = time.perf_counter()
        linhas_lidas = 0

        # read_only + values_only reduz bastante tempo/memória em planilhas grandes
        workbook = openpyxl.load_workbook(self.caminho, read_only=True, data_only=True)
        try:
            worksheet = workbook.active

            # Extrair cabeçalhos (linha 1)
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return

            cabecalhos = [str(h).strip() if h is not None else None for h in header_row]
            logger.debug(f"Cabeçalhos encontrados: {cabecalhos}")

            # Manter apenas colunas necessárias para montar Alocacao (evita construir dict gigante)
            cabecalhos_necessarios = {
                "CHAVE",
                "Placa",
                "Identificador da rota",
                "Identificador",
                "Pedido",
                "NF",
                "Cliente",
                "Cidade",
                "Tipo cliente",
                "Bairro",
                "Endereço",
                "Cep",
                "Valor total pedido",
                "Qtd. caixas",
                "Peso bruto pedido",
                "Distância calculado",
                "Código cliente",
            }

            indices: list[int] = [i for i, h in enumerate(cabecalhos) if h and str(h) in cabecalhos_necessarios]
            # fallback: se não detectou cabeçalhos esperados, usa todas as colunas não-vazias
            if not indices:
                indices = [i for i, h in enumerate(cabecalhos) if h]

            rota_atual: Optional[str] = None
            aguardando_nome_rota = False
            quebra_id_atual: Optional[str] = None

            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row:
                    continue

                col_a = row[0] if len(row) > 0 else None
                col_a_str = str(col_a).strip() if col_a is not None else ""

                # Linhas de quebra: "Quebra Por: ..." (não são dados)
                if col_a_str.startswith("Quebra"):
                    logger.debug(f"Ignorando linha {idx} (quebra): {col_a_str}")
                    # Extrair um ID numérico desta linha (ex.: "Quebra Por: ... (608825)" ou "Quebra Por: 608825")
                    m = re.search(r"(\d{3,})", col_a_str)
                    quebra_id_atual = m.group(1) if m else None
                    aguardando_nome_rota = True
                    continue

                # Próxima linha após "Quebra" é o nome da rota.
                # Observação: em algumas planilhas essa linha também já contém dados (ex.: CHAVE).
                # Então capturamos a rota e deixamos o fluxo seguir; o filtro por CHAVE decide se é dado.
                if aguardando_nome_rota:
                    if col_a_str:
                        rota_atual = col_a_str
                        logger.debug(f"Rota detectada na linha {idx}: {rota_atual}")
                    aguardando_nome_rota = False

                # Pular linhas de totalização (ex.: "Total: 13987.68")
                if any(isinstance(v, str) and v.strip().startswith("Total:") for v in row if v is not None):
                    logger.debug(f"Ignorando linha {idx} (totalização)")
                    continue

                linha_dados: Dict = {}
                for col_idx in indices:
                    if col_idx >= len(row):
                        continue
                    cabecalho = cabecalhos[col_idx]
                    if cabecalho:
                        linha_dados[cabecalho] = row[col_idx]

                # Preencher rota atual quando a coluna de identificador estiver vazia
                if rota_atual:
                    for chave_rota in ("Identificador da rota", "Identificador"):
                        if chave_rota in linha_dados and not linha_dados.get(chave_rota):
                            linha_dados[chave_rota] = rota_atual

                # Propagar ID único da quebra para as linhas de dados
                if quebra_id_atual:
                    linha_dados["ID_QUEBRA"] = quebra_id_atual

                # Adicionar somente linhas que parecem dados (precisa ter CHAVE)
                chave = linha_dados.get("CHAVE")
                if chave is None or str(chave).strip() == "":
                    continue

                linhas_lidas += 1
                yield linha_dados
        finally:
            try:
                workbook.close()
            except Exception:
                pass

            # Log no finally para contar inclusive em consumo parcial
            logger.info(
                f"Lidas {linhas_lidas} linhas de dados da planilha em {time.perf_counter() - t0:.2f}s"
            )

    def ler_dados_brutos(self) -> List[Dict]:
        """Lê os dados brutos da planilha (lista completa)."""
        try:
            return list(self.iter_dados_brutos())
        except Exception as e:
            logger.error(f"Erro ao ler planilha: {e}")
            raise
    
    def ler_alocacoes(self) -> List[Alocacao]:
        """
        Lê a planilha e retorna lista de objetos Alocacao.
        
        Returns:
            Lista de Alocacao
        """
        alocacoes: List[Alocacao] = []

        for dados in self.iter_dados_brutos():
            try:
                alocacao = Alocacao.from_dict(dados)
                alocacoes.append(alocacao)
            except Exception as e:
                logger.warning(f"Erro ao processar linha: {e}")
                continue
        
        return alocacoes
